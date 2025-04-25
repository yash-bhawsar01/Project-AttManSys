import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Get the absolute path to the images folder inside 'facial_attendance' on Desktop
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "facial_attendance", "images")

# Roll numbers and corresponding student names
student_data = {
    150: "Yash",
    151: "Yash Salvi",
    149: "Vivek DeDwa"
}

# Load images for all students
images = []
classnames = []
for roll_no, name in student_data.items():
    image_path = os.path.join(desktop_path, f"{roll_no}.jpg")  # Images named by roll numbers
    curImg = cv2.imread(image_path)
    if curImg is None:
        print(f"Image {roll_no}.jpg could not be read.")
        continue
    images.append(curImg)
    classnames.append(name)

# Encode faces function
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        face_encodings = face_recognition.face_encodings(img)
        
        if face_encodings:  # Check if any face encodings are found
            encodeList.append(face_encodings[0])
        else:
            print("No face detected in one of the images.")
    return encodeList

# Save attendance to Excel function
def markAttendance(roll_no, name):
    # Get the path to save the attendance file in the same folder as main.py
    attendance_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'attendance.xlsx')
    
    # Load the existing workbook and sheet
    workbook = load_workbook(attendance_file)
    sheet = workbook.active

    # Check if roll number already exists in attendance
    roll_numbers = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row)]
    
    if roll_no in roll_numbers:
        # Find the row where the student's roll number is located
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
            if row[0].value == roll_no:
                present_column = get_column_letter(3)  # Assuming "Present" is in column 3 (C)
                row_num = row[0].row
                # Mark "Present" with a tick mark (✓)
                sheet[f'{present_column}{row_num}'] = '✓'
                # Add the current time and date
                now = datetime.now()
                timeStr = now.strftime('%H:%M:%S')
                dateStr = now.strftime('%Y-%m-%d')
                sheet[f'D{row_num}'] = timeStr
                sheet[f'E{row_num}'] = dateStr
                workbook.save(attendance_file)

# Prepare encodings for all students
encodeListKnown = findEncodings(images)
print('Encoding Complete')

# Start webcam
cap = cv2.VideoCapture(0)

while True:
    success, img = cap.read()
    imgSmall = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
    imgSmall = cv2.cvtColor(imgSmall, cv2.COLOR_BGR2RGB)

    # Find faces and encodings in the current frame
    facesCurFrame = face_recognition.face_locations(imgSmall)
    encodesCurFrame = face_recognition.face_encodings(imgSmall, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = classnames[matchIndex]
            roll_no = list(student_data.keys())[matchIndex]  # Get roll number from the index
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)

            # Display name with a background color box around it
            font = cv2.FONT_HERSHEY_SIMPLEX
            text = f"{name} ({roll_no})"
            (w, h), _ = cv2.getTextSize(text, font, 0.8, 2)
            cv2.rectangle(img, (x1, y1 - 35), (x1 + w, y1), (0, 255, 0), -1)  # Background box
            cv2.putText(img, text, (x1 + 6, y1 - 6), font, 0.8, (255, 255, 255), 2)
            markAttendance(roll_no, name)

    cv2.imshow('Webcam - Press q to quit', img)
    if cv2.waitKey(1) == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
